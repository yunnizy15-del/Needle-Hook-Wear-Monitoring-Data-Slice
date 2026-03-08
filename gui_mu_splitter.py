#!/usr/bin/env python3
from __future__ import annotations

import queue
import sys
import threading
import traceback
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook

from split_mu_by_tlife import ProcessStats, process_xlsx


class MuSplitterApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self._set_app_icon()
        self.title("针钩摩擦系数数据切片工具")
        self.geometry("900x620")
        self.minsize(760, 520)

        self.input_var = tk.StringVar()
        self.tlife_var = tk.StringVar()
        self.slice_seconds_var = tk.StringVar(value="5")
        self.drop_minutes_var = tk.StringVar(value="30")
        self.drop_initial_hours_var = tk.StringVar(value="1")
        self.output_var = tk.StringVar(value=str(Path.cwd()))
        self.status_var = tk.StringVar(value="就绪")

        default_input = Path.cwd() / "data.xlsx"
        if default_input.exists():
            self.input_var.set(str(default_input))

        self._queue: queue.Queue[tuple[str, object]] = queue.Queue()
        self._sheet_names: list[str] = []
        self._worker: threading.Thread | None = None

        self._build_ui()
        if default_input.exists():
            self._load_sheets(show_popup=False)
        self.after(150, self._poll_queue)

    def _set_app_icon(self) -> None:
        # Prefer packaged resource path when running from PyInstaller.
        base_dir = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
        icon_path = base_dir / "app.ico"
        if icon_path.exists():
            try:
                self.iconbitmap(str(icon_path))
            except Exception:
                pass

    def _build_ui(self) -> None:
        style = ttk.Style()
        style.configure("Primary.TButton", font=("Microsoft YaHei UI", 10, "bold"))

        root = ttk.Frame(self, padding=14)
        root.pack(fill=tk.BOTH, expand=True)
        root.columnconfigure(1, weight=1)

        ttk.Label(root, text="Excel 文件：", font=("Microsoft YaHei UI", 10)).grid(
            row=0, column=0, sticky="w", pady=4
        )
        self.input_entry = ttk.Entry(root, textvariable=self.input_var)
        self.input_entry.grid(row=0, column=1, sticky="ew", padx=(0, 8), pady=4)
        self.input_btn = ttk.Button(root, text="浏览...", command=self._choose_input)
        self.input_btn.grid(row=0, column=2, pady=4)

        ttk.Label(root, text="选择工作表：", font=("Microsoft YaHei UI", 10)).grid(
            row=1, column=0, sticky="w", pady=4
        )
        sheet_frame = ttk.Frame(root)
        sheet_frame.grid(row=1, column=1, sticky="ew", padx=(0, 8), pady=4)
        sheet_frame.columnconfigure(0, weight=1)
        sheet_frame.rowconfigure(0, weight=1)
        self.sheet_listbox = tk.Listbox(
            sheet_frame,
            selectmode=tk.MULTIPLE,
            exportselection=False,
            height=5,
            font=("Microsoft YaHei UI", 10),
        )
        self.sheet_listbox.grid(row=0, column=0, sticky="ew")
        self.sheet_scrollbar = ttk.Scrollbar(sheet_frame, orient="vertical", command=self.sheet_listbox.yview)
        self.sheet_scrollbar.grid(row=0, column=1, sticky="ns")
        self.sheet_listbox.configure(yscrollcommand=self.sheet_scrollbar.set)
        self.refresh_sheet_btn = ttk.Button(root, text="刷新工作表", command=self._on_refresh_sheets)
        self.refresh_sheet_btn.grid(row=1, column=2, pady=4)

        ttk.Label(root, text="tlife（秒）：", font=("Microsoft YaHei UI", 10)).grid(
            row=2, column=0, sticky="w", pady=4
        )
        self.tlife_entry = ttk.Entry(root, textvariable=self.tlife_var)
        self.tlife_entry.grid(row=2, column=1, sticky="ew", padx=(0, 8), pady=4)

        ttk.Label(root, text="切片秒数：", font=("Microsoft YaHei UI", 10)).grid(
            row=3, column=0, sticky="w", pady=4
        )
        self.slice_entry = ttk.Entry(root, textvariable=self.slice_seconds_var)
        self.slice_entry.grid(row=3, column=1, sticky="ew", padx=(0, 8), pady=4)

        ttk.Label(root, text="剔除窗口（分钟）：", font=("Microsoft YaHei UI", 10)).grid(
            row=4, column=0, sticky="w", pady=4
        )
        self.drop_entry = ttk.Entry(root, textvariable=self.drop_minutes_var)
        self.drop_entry.grid(row=4, column=1, sticky="ew", padx=(0, 8), pady=4)

        ttk.Label(root, text="丢弃前 x 小时：", font=("Microsoft YaHei UI", 10)).grid(
            row=5, column=0, sticky="w", pady=4
        )
        self.drop_initial_entry = ttk.Entry(root, textvariable=self.drop_initial_hours_var)
        self.drop_initial_entry.grid(row=5, column=1, sticky="ew", padx=(0, 8), pady=4)

        ttk.Label(root, text="输出目录：", font=("Microsoft YaHei UI", 10)).grid(
            row=6, column=0, sticky="w", pady=4
        )
        self.output_entry = ttk.Entry(root, textvariable=self.output_var)
        self.output_entry.grid(row=6, column=1, sticky="ew", padx=(0, 8), pady=4)
        self.output_btn = ttk.Button(root, text="浏览...", command=self._choose_output)
        self.output_btn.grid(row=6, column=2, pady=4)

        hint = (
            "规则：t < tlife-30分钟 记为有效；t > tlife+30分钟 记为失效；中间区间丢弃。\n"
            "输出：在输出目录下生成 valid / invalid 两个文件夹，按序号保存 CSV。"
        )
        ttk.Label(root, text=hint, foreground="#444").grid(
            row=7, column=0, columnspan=3, sticky="w", pady=(6, 8)
        )

        action_frame = ttk.Frame(root)
        action_frame.grid(row=8, column=0, columnspan=3, sticky="ew")
        action_frame.columnconfigure(0, weight=1)

        self.start_btn = ttk.Button(
            action_frame,
            text="开始处理",
            style="Primary.TButton",
            command=self._start_processing,
        )
        self.start_btn.grid(row=0, column=0, sticky="w")

        self.progress = ttk.Progressbar(action_frame, mode="indeterminate", length=220)
        self.progress.grid(row=0, column=1, sticky="e")

        ttk.Label(root, textvariable=self.status_var, foreground="#0b62a4").grid(
            row=9, column=0, columnspan=3, sticky="w", pady=(8, 6)
        )

        log_frame = ttk.Frame(root)
        log_frame.grid(row=10, column=0, columnspan=3, sticky="nsew")
        root.rowconfigure(10, weight=1)
        log_frame.rowconfigure(0, weight=1)
        log_frame.columnconfigure(0, weight=1)

        self.log_text = tk.Text(log_frame, height=16, wrap="word", font=("Consolas", 10))
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.log_text.configure(yscrollcommand=scrollbar.set)

    def _choose_input(self) -> None:
        path = filedialog.askopenfilename(
            title="选择 xlsx 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if path:
            self.input_var.set(path)
            self._load_sheets(show_popup=True)

    def _choose_output(self) -> None:
        path = filedialog.askdirectory(title="选择输出目录")
        if path:
            self.output_var.set(path)

    def _append_log(self, text: str) -> None:
        self.log_text.insert(tk.END, text + "\n")
        self.log_text.see(tk.END)

    def _set_running(self, running: bool) -> None:
        state = tk.DISABLED if running else tk.NORMAL
        for widget in [
            self.input_entry,
            self.input_btn,
            self.refresh_sheet_btn,
            self.sheet_listbox,
            self.tlife_entry,
            self.slice_entry,
            self.drop_entry,
            self.drop_initial_entry,
            self.output_entry,
            self.output_btn,
            self.start_btn,
        ]:
            widget.configure(state=state)

        if running:
            self.progress.start(12)
        else:
            self.progress.stop()

    def _on_refresh_sheets(self) -> None:
        self._load_sheets(show_popup=True)

    def _selected_sheets(self) -> list[str]:
        selected = [self._sheet_names[i] for i in self.sheet_listbox.curselection() if i < len(self._sheet_names)]
        return selected

    def _load_sheets(self, show_popup: bool) -> None:
        raw_path = self.input_var.get().strip().strip('"')
        if not raw_path:
            if show_popup:
                messagebox.showerror("提示", "请先选择 Excel 文件。")
            return

        input_path = Path(raw_path).expanduser()
        if not input_path.exists():
            self._sheet_names = []
            self.sheet_listbox.delete(0, tk.END)
            if show_popup:
                messagebox.showerror("提示", "Excel 文件不存在。")
            return

        try:
            wb = load_workbook(input_path, read_only=True, data_only=True)
            try:
                sheet_names = list(wb.sheetnames)
            finally:
                wb.close()
            if not sheet_names:
                raise ValueError("文件中没有可用工作表。")
        except Exception as exc:
            self._sheet_names = []
            self.sheet_listbox.delete(0, tk.END)
            self._append_log(f"读取工作表失败: {exc}")
            if show_popup:
                messagebox.showerror("读取失败", str(exc))
            return

        self._sheet_names = sheet_names
        self.sheet_listbox.delete(0, tk.END)
        for s in sheet_names:
            self.sheet_listbox.insert(tk.END, s)
        default_count = min(2, len(sheet_names))
        for i in range(default_count):
            self.sheet_listbox.selection_set(i)

        self.status_var.set(f"已加载 {len(sheet_names)} 个工作表，默认已选前 {default_count} 个")
        self._append_log(f"可选工作表: {', '.join(sheet_names)}")

    def _start_processing(self) -> None:
        if self._worker is not None and self._worker.is_alive():
            return

        try:
            input_path = Path(self.input_var.get().strip().strip('"')).expanduser()
            if not input_path.exists():
                raise ValueError("Excel 文件不存在。")
            tlife = float(self.tlife_var.get().strip())
            slice_seconds = float(self.slice_seconds_var.get().strip())
            drop_minutes = float(self.drop_minutes_var.get().strip())
            drop_initial_hours = float(self.drop_initial_hours_var.get().strip())
            if slice_seconds <= 0:
                raise ValueError("切片秒数必须大于 0。")
            if drop_minutes < 0:
                raise ValueError("剔除窗口分钟数不能小于 0。")
            if drop_initial_hours < 0:
                raise ValueError("丢弃前x小时不能小于 0。")
            output_dir_text = self.output_var.get().strip().strip('"')
            if not output_dir_text:
                output_dir_text = "."
            output_dir = Path(output_dir_text).expanduser()
            sheets = self._selected_sheets()
            if not sheets:
                raise ValueError("请至少选择一个工作表。")
        except Exception as exc:
            messagebox.showerror("参数错误", str(exc))
            return

        self.status_var.set("处理中，请稍候...")
        self._append_log("开始处理...")
        self._append_log(f"输入文件: {input_path}")
        self._append_log(f"工作表: {', '.join(sheets)}")
        self._append_log(
            f"tlife: {tlife} 秒, 切片: {slice_seconds} 秒, "
            f"剔除窗口: {drop_minutes} 分钟, 丢弃前: {drop_initial_hours} 小时"
        )
        self._set_running(True)

        self._worker = threading.Thread(
            target=self._run_worker,
            args=(
                input_path,
                tlife,
                slice_seconds,
                drop_minutes,
                drop_initial_hours,
                sheets,
                output_dir,
            ),
            daemon=True,
        )
        self._worker.start()

    def _run_worker(
        self,
        input_path: Path,
        tlife: float,
        slice_seconds: float,
        drop_minutes: float,
        drop_initial_hours: float,
        sheets: list[str],
        output_dir: Path,
    ) -> None:
        try:
            all_stats: list[ProcessStats] = []
            total = len(sheets)
            valid_seq = 0
            invalid_seq = 0
            for idx, sheet_name in enumerate(sheets, start=1):
                self._queue.put(("sheet_start", (idx, total, sheet_name, output_dir)))
                stats = process_xlsx(
                    input_path=input_path,
                    tlife=tlife,
                    slice_seconds=slice_seconds,
                    drop_minutes=drop_minutes,
                    drop_initial_hours=drop_initial_hours,
                    sheet=sheet_name,
                    output_dir=output_dir,
                    clear_output=(idx == 1),
                    valid_start_seq=valid_seq,
                    invalid_start_seq=invalid_seq,
                    progress_callback=lambda n, sn=sheet_name, i=idx, t=total: self._queue.put(
                        ("progress", (i, t, sn, n))
                    ),
                )
                valid_seq = stats.valid_last_seq
                invalid_seq = stats.invalid_last_seq
                all_stats.append(stats)
                self._queue.put(("sheet_done", (idx, total, stats)))
            self._queue.put(("done", all_stats))
        except Exception as exc:
            tb = traceback.format_exc()
            self._queue.put(("error", (exc, tb)))

    def _format_done(self, stats: ProcessStats) -> str:
        return (
            f"处理完成\n"
            f"Sheet: {stats.sheet_name}\n"
            f"丢弃区间: [{stats.valid_upper_bound:.3f}, {stats.invalid_lower_bound:.3f}] 秒\n"
            f"有效数据: {stats.valid_rows} 行, 文件: {stats.valid_files}\n"
            f"失效数据: {stats.invalid_rows} 行, 文件: {stats.invalid_files}\n"
            f"丢弃: {stats.dropped_rows} 行 "
            f"(前置丢弃: {stats.dropped_initial_rows}, tlife窗口丢弃: {stats.dropped_tlife_rows}), "
            f"跳过: {stats.skipped_rows} 行, 扫描: {stats.scanned_rows} 行\n"
            f"输出目录: {stats.output_dir}"
        )

    def _format_summary(self, all_stats: list[ProcessStats]) -> str:
        valid_rows = sum(s.valid_rows for s in all_stats)
        invalid_rows = sum(s.invalid_rows for s in all_stats)
        dropped_rows = sum(s.dropped_rows for s in all_stats)
        dropped_initial_rows = sum(s.dropped_initial_rows for s in all_stats)
        dropped_tlife_rows = sum(s.dropped_tlife_rows for s in all_stats)
        skipped_rows = sum(s.skipped_rows for s in all_stats)
        scanned_rows = sum(s.scanned_rows for s in all_stats)
        valid_files_total = all_stats[-1].valid_last_seq if all_stats else 0
        invalid_files_total = all_stats[-1].invalid_last_seq if all_stats else 0
        return (
            f"全部工作表处理完成（{len(all_stats)} 个）\n"
            f"有效数据总计: {valid_rows} 行\n"
            f"有效切片文件总计: {valid_files_total} 个\n"
            f"失效数据总计: {invalid_rows} 行\n"
            f"失效切片文件总计: {invalid_files_total} 个\n"
            f"丢弃总计: {dropped_rows} 行 "
            f"(前置丢弃: {dropped_initial_rows}, tlife窗口丢弃: {dropped_tlife_rows}), "
            f"跳过总计: {skipped_rows} 行, 扫描总计: {scanned_rows} 行"
        )

    def _poll_queue(self) -> None:
        while True:
            try:
                event, payload = self._queue.get_nowait()
            except queue.Empty:
                break

            if event == "sheet_start":
                idx, total, sheet_name, out_dir = payload
                self.status_var.set(f"处理中... ({idx}/{total}) {sheet_name}")
                self._append_log(f"[{idx}/{total}] 开始处理工作表: {sheet_name}")
                self._append_log(f"输出目录: {out_dir}")
            elif event == "progress":
                idx, total, sheet_name, rows = payload
                self.status_var.set(f"处理中... ({idx}/{total}) {sheet_name} 已扫描 {rows} 行")
                self._append_log(f"[{idx}/{total}] {sheet_name}: 已扫描 {rows} 行")
            elif event == "sheet_done":
                idx, total, stats = payload
                self._append_log(f"[{idx}/{total}] {stats.sheet_name} 已完成")
                self._append_log(self._format_done(stats))
            elif event == "done":
                all_stats = payload
                self._set_running(False)
                self.status_var.set("处理完成")
                summary = self._format_summary(all_stats)
                self._append_log(summary)
                messagebox.showinfo("完成", "处理完成，结果已输出。")
            elif event == "error":
                exc, tb = payload
                self._set_running(False)
                self.status_var.set("处理失败")
                self._append_log(f"错误: {exc}")
                self._append_log(tb)
                messagebox.showerror("处理失败", str(exc))

        self.after(150, self._poll_queue)


def main() -> None:
    app = MuSplitterApp()
    app.mainloop()


if __name__ == "__main__":
    main()
