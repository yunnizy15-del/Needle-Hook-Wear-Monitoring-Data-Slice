#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable, Optional

from openpyxl import load_workbook


SECONDS_PER_MINUTE = 60.0


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def find_column_index(headers: Iterable[object], keywords: list[str]) -> int:
    normalized_headers = [normalize_text(h) for h in headers]
    for i, header in enumerate(normalized_headers):
        if any(k in header for k in keywords):
            return i
    raise ValueError(f"Column not found. Keywords: {keywords}")


def to_float(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip()
        if not text:
            return None
        try:
            number = float(text)
        except ValueError:
            return None
    if not math.isfinite(number):
        return None
    return number


def choose_xlsx_file() -> Optional[Path]:
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename(
            title="Choose xlsx file",
            filetypes=[("Excel file", "*.xlsx")],
        )
        root.destroy()
        if not file_path:
            return None
        return Path(file_path)
    except Exception:
        return None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Split mu_true by tlife and save time slices into valid/invalid folders."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help="Path to xlsx file; if omitted, try file dialog.",
    )
    parser.add_argument(
        "--tlife",
        type=float,
        default=None,
        help="tlife in seconds.",
    )
    parser.add_argument(
        "--slice-seconds",
        type=float,
        default=5.0,
        help="Slice length in seconds. Default: 5.",
    )
    parser.add_argument(
        "--drop-minutes",
        type=float,
        default=30.0,
        help="Drop window on each side of tlife in minutes. Default: 30.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name; default is the first sheet.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("."),
        help="Output directory. Default: current directory.",
    )
    return parser.parse_args()


def ensure_output_dir(folder: Path) -> None:
    folder.mkdir(parents=True, exist_ok=True)
    for old_csv in folder.glob("*.csv"):
        old_csv.unlink()


@dataclass
class SliceWriter:
    folder: Path
    slice_seconds: float
    file_seq: int = 0
    anchor_time: Optional[float] = None
    current_window: Optional[int] = None
    current_file = None
    current_writer: Optional[csv.writer] = None
    total_rows: int = 0

    def _open_new_slice(self, window_index: int) -> None:
        self.close()
        self.file_seq += 1
        self.current_window = window_index
        file_path = self.folder / f"{self.file_seq:06d}.csv"
        self.current_file = file_path.open("w", newline="", encoding="utf-8-sig")
        self.current_writer = csv.writer(self.current_file)
        self.current_writer.writerow(["t_s", "mu_true"])

    def write(self, time_s: float, mu_true: float) -> None:
        if self.anchor_time is None:
            self.anchor_time = time_s
        window_index = int((time_s - self.anchor_time) // self.slice_seconds)
        if self.current_window != window_index or self.current_writer is None:
            self._open_new_slice(window_index)
        self.current_writer.writerow([f"{time_s:.6f}", f"{mu_true:.12g}"])
        self.total_rows += 1

    def close(self) -> None:
        if self.current_file is not None:
            self.current_file.close()
        self.current_file = None
        self.current_writer = None


@dataclass
class ProcessStats:
    input_file: Path
    output_dir: Path
    sheet_name: str
    tlife: float
    valid_upper_bound: float
    invalid_lower_bound: float
    scanned_rows: int
    skipped_rows: int
    dropped_rows: int
    valid_rows: int
    invalid_rows: int
    valid_files: int
    invalid_files: int


def process_xlsx(
    input_path: Path,
    tlife: float,
    slice_seconds: float = 5.0,
    drop_minutes: float = 30.0,
    sheet: Optional[str] = None,
    output_dir: Path = Path("."),
    progress_every: int = 200000,
    progress_callback: Optional[Callable[[int], None]] = None,
) -> ProcessStats:
    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_path}")
    if slice_seconds <= 0:
        raise ValueError("--slice-seconds must be > 0.")
    if drop_minutes < 0:
        raise ValueError("--drop-minutes must be >= 0.")
    if progress_every <= 0:
        raise ValueError("--progress_every must be > 0.")

    drop_seconds = drop_minutes * SECONDS_PER_MINUTE
    valid_upper_bound = tlife - drop_seconds
    invalid_lower_bound = tlife + drop_seconds

    output_root = output_dir.resolve()
    valid_dir = output_root / "valid"
    invalid_dir = output_root / "invalid"
    ensure_output_dir(valid_dir)
    ensure_output_dir(invalid_dir)

    valid_writer = SliceWriter(folder=valid_dir, slice_seconds=slice_seconds)
    invalid_writer = SliceWriter(folder=invalid_dir, slice_seconds=slice_seconds)

    wb = load_workbook(input_path, read_only=True, data_only=True)
    selected_sheet_name = ""
    total_rows = 0
    dropped_rows = 0
    skipped_rows = 0
    try:
        if sheet:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet does not exist: {sheet}")
            ws = wb[sheet]
        else:
            ws = wb[wb.sheetnames[0]]
        selected_sheet_name = ws.title

        rows = ws.iter_rows(values_only=True)
        headers = next(rows)
        if headers is None:
            raise ValueError("Header row is empty.")

        time_col = find_column_index(headers, ["t_s", "time"])
        mu_col = find_column_index(headers, ["mu_true"])

        for total_rows, row in enumerate(rows, start=1):
            if row is None:
                skipped_rows += 1
                continue
            time_s = to_float(row[time_col] if time_col < len(row) else None)
            mu_true = to_float(row[mu_col] if mu_col < len(row) else None)
            if time_s is None or mu_true is None:
                skipped_rows += 1
                continue

            if time_s < valid_upper_bound:
                valid_writer.write(time_s, mu_true)
            elif time_s > invalid_lower_bound:
                invalid_writer.write(time_s, mu_true)
            else:
                dropped_rows += 1

            if total_rows % progress_every == 0 and progress_callback is not None:
                progress_callback(total_rows)
    finally:
        valid_writer.close()
        invalid_writer.close()
        wb.close()

    return ProcessStats(
        input_file=input_path.resolve(),
        output_dir=output_root,
        sheet_name=selected_sheet_name,
        tlife=tlife,
        valid_upper_bound=valid_upper_bound,
        invalid_lower_bound=invalid_lower_bound,
        scanned_rows=total_rows,
        skipped_rows=skipped_rows,
        dropped_rows=dropped_rows,
        valid_rows=valid_writer.total_rows,
        invalid_rows=invalid_writer.total_rows,
        valid_files=valid_writer.file_seq,
        invalid_files=invalid_writer.file_seq,
    )


def prompt_float(message: str) -> float:
    while True:
        raw = input(message).strip()
        try:
            value = float(raw)
            if math.isfinite(value):
                return value
        except ValueError:
            pass
        print("Invalid input. Please enter a numeric value.")


def main() -> None:
    args = parse_args()

    input_path = args.input
    if input_path is None:
        input_path = choose_xlsx_file()
    if input_path is None:
        raw_path = input("Please enter xlsx path: ").strip().strip('"')
        input_path = Path(raw_path)
    tlife = args.tlife if args.tlife is not None else prompt_float("Please enter tlife (seconds): ")
    stats = process_xlsx(
        input_path=input_path,
        tlife=tlife,
        slice_seconds=args.slice_seconds,
        drop_minutes=args.drop_minutes,
        sheet=args.sheet,
        output_dir=args.output_dir,
        progress_callback=lambda n: print(f"Processed {n} rows..."),
    )

    print("Done.")
    print(f"Input file: {stats.input_file}")
    print(f"Sheet: {stats.sheet_name}")
    print(
        f"tlife: {stats.tlife} s, dropped interval: "
        f"[{stats.valid_upper_bound}, {stats.invalid_lower_bound}] s"
    )
    print(f"Valid rows: {stats.valid_rows}, slice files: {stats.valid_files}")
    print(f"Invalid rows: {stats.invalid_rows}, slice files: {stats.invalid_files}")
    print(
        f"Dropped rows: {stats.dropped_rows}, skipped rows: {stats.skipped_rows}, "
        f"scanned rows: {stats.scanned_rows}"
    )
    print(f"Output directory: {stats.output_dir}")


if __name__ == "__main__":
    main()
